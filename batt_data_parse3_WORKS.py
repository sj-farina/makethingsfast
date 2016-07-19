import openpyxl

#DEFINES
INDEX_COL = 1
CUR_COL = 7
VOL_COL = 8
CAP_MAH_COL = 10
CAP_PWR_COL = 12
TEMP_COL = 22


#First, load the workbook and sheet you want
#wb should be in same directory as code
###################################################################################################### TODO UPDATE THIS
ws = openpyxl.load_workbook('testbook.xlsx').get_sheet_by_name('Sheet1')


#First get the capacity and power
#scan voltage
#if <5, record value of Discharge Capacity and energy, along with temp
cel = ws['H1']
i = 2
charging = True
while True:
	cel = ws.cell(row= i,column= VOL_COL)
	#print(cel.value)
	if cel.value == None:
		break
	if cel.value < 5 and ws.cell(row= i,column= CAP_MAH_COL).value != 0:
		charging = False if ws.cell(row= i,column= CUR_COL).value <= 0 else True
		if charging == False:
			print (ws.cell(row= i,column= INDEX_COL).value, cel.value, ws.cell(row= i,column= CAP_MAH_COL).value, ws.cell(row= i,column= CAP_PWR_COL).value, ws.cell(row= i,column= TEMP_COL).value)
	i +=1

print('______________________')


#next we need to find the transitions from 0 current to not
cur_cell = ws['A1']
last_cell = ws['A1']
last_jump = 1
cur_jump = 1
###################################################################################################### TODO UPDATE THIS
#should fix this, make the subtraction automatic
v_target = [8.2-0.001, 8.0-0.001, 7.8-0.001, 7.6-0.001, 7.4-0.001, 0]


i = 2
j = 0
new_dataset = True
while True:
	#iterate through the cells, saving current and last
	last_cell = cur_cell
	cur_cell = ws.cell(row= i,column= CUR_COL)
	#if you run out of data, break
	if cur_cell.value == None:
		break
	#check for new tests
	if ws.cell(row= i,column= VOL_COL).value < 5:
		print('______________________')
		new_dataset = True
		j = 0
	#look at current, every falling edge, record position, check values
	if last_cell.value == 0 and cur_cell.value < 0:
		last_jump = cur_jump
		cur_jump = i
		if new_dataset == True:
			print (ws.cell(row= cur_jump - 1,column= INDEX_COL).value, ws.cell(row= cur_jump - 1,column= VOL_COL).value)
			print (ws.cell(row= cur_jump,column= INDEX_COL).value, ws.cell(row= cur_jump,column= VOL_COL).value)
			new_dataset = False
		elif ws.cell(row= cur_jump - 1, column= VOL_COL).value < v_target[j] < ws.cell(row= last_jump - 1, column= VOL_COL).value :
			print('')
			print (ws.cell(row= last_jump - 1,column= INDEX_COL).value, ws.cell(row= last_jump - 1,column= VOL_COL).value)
			print (ws.cell(row= last_jump,column= INDEX_COL).value, ws.cell(row= last_jump,column= VOL_COL).value)
			j += 1
	i += 1
		
