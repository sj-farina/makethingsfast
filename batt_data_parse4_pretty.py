##############################################################################
# batt_data_parse4.py
# Janey Farina
# Last update: 7-18-16
#
# Quickly processes data from the battery tester
#
# Finds capacity at end of a given measurement
# Finds voltage drop when discharging begins at given starting voltage
#
# Automatically recognizes multiple testing temperatures in single data set
# 
# To set starting voltages, change v_target
# To change file, edit data_in
#
# To double check by viewing line numbers, toggle print statements
##############################################################################


import openpyxl


# DEFINES
INDEX_COL = 1
CUR_COL = 7
VOL_COL = 8
CAP_MAH_COL = 10
CAP_PWR_COL = 12
TEMP_COL = 22
# Include negative sign here for discharging
DISCHARGE_CUR = -4





###############
# Part 0: Setup
# Change file and search values
###############

# File must be in same working directory, check sheet name
data_in = openpyxl.load_workbook('testbook.xlsx').get_sheet_by_name('Sheet1')
# Enter numbers here to change search values, leave zero at the end
v_target = [4.1, 4.0, 3.9, 3.8, 3.7, 0]
CUTOFF = 2.5


###############
# Part 1
# Find capacity
###############

# Search for voltage under CUTOFF to indicate cycle end
cur_cell = data_in['A1']
i = 2
print("mAh capacity, Power capacity")
while True:
	cur_cell = data_in.cell(row = i, column = VOL_COL)
	if cur_cell.value == None:
		break
	if (cur_cell.value < CUTOFF and data_in.cell(row = i, column = CAP_MAH_COL).value != 0 
	and data_in.cell(row = i, column = CUR_COL).value <= DISCHARGE_CUR + 1):
		print ("Temperature = ", round(data_in.cell(row = i, column = TEMP_COL).value))
		# Toggle next two lines to print cell index
		print (data_in.cell(row = i, column = CAP_MAH_COL).value, 
			data_in.cell(row = i, column = CAP_PWR_COL).value)
		#print (data_in.cell(row = i, column = INDEX_COL).value, 
		#	data_in.cell(row = i, column = CAP_MAH_COL).value, 
		#	data_in.cell(row = i, column = CAP_PWR_COL).value)
	i +=1

print('________________________')




###############
# Part 2
# Find the voltage drops as discharge starts
###############

# Initial values do not matter as long as not null
cur_cell = data_in['A1']
last_cell = data_in['A1']
last_jump = 1
cur_jump = 1

# Automatically subtracts .001 for rounding purposes 
for num in range(len(v_target)):
	v_target[num] = v_target[num] - 0.001

# Generic counters for looping
i = 2
j = 0

new_dataset = True
while True:
	# Iterate through the cells, saving current and previous 
	last_cell = cur_cell
	cur_cell = data_in.cell(row = i, column = CUR_COL)
	# If you run out of data, break
	if cur_cell.value == None:
		break
	# Check for new tests
	if data_in.cell(row = i, column = VOL_COL).value < CUTOFF:
		print('______________________')
		new_dataset = True
		j = 0
	# Look at current, every falling edge, record position, check values
	if last_cell.value == 0 and cur_cell.value < DISCHARGE_CUR + 1:
		last_jump = cur_jump
		cur_jump = i
		if new_dataset == True:
			print('Temperature = ', round(data_in.cell(row = cur_jump - 1, 
				column = TEMP_COL).value))
			print('Voltage at Start')
			# Toggle next four lines to print cell index
			print (data_in.cell(row = cur_jump - 1, column = VOL_COL).value)
			print (data_in.cell(row = cur_jump, column = VOL_COL).value)
			#print (data_in.cell(row = cur_jump - 1, column = INDEX_COL).value, 
			#data_in.cell(row = cur_jump - 1, column = VOL_COL).value)
			#print (data_in.cell(row = cur_jump, column = INDEX_COL).value, 
			#data_in.cell(row = cur_jump, column = VOL_COL).value)
			new_dataset = False

		elif (data_in.cell(row = cur_jump - 1,  column = VOL_COL).value < 
		v_target[j]	< data_in.cell(row = last_jump - 1,  column = VOL_COL).value):
			print('Voltage = ', v_target[j] + .001)
			# Toggle next four lines to print cell index
			print (data_in.cell(row = last_jump - 1, column = VOL_COL).value)
			print (data_in.cell(row = last_jump, column = VOL_COL).value)			
			#print (data_in.cell(row = last_jump - 1, column = INDEX_COL).value, 
			#data_in.cell(row = last_jump - 1, column = VOL_COL).value)
			#print (data_in.cell(row = last_jump, column = INDEX_COL).value, 
			#data_in.cell(row = last_jump, column = VOL_COL).value)
			j += 1
	i += 1
		
